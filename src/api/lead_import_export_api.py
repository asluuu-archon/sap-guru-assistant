"""
Lead Import & Export API

Import:
  POST /leads/import/preview  — parse CSV/Excel, return preview rows + column mapping suggestions
  POST /leads/import/confirm  — actually insert the rows into the leads table

Export:
  GET  /leads/export/csv      — download all leads (or filtered) as CSV
  GET  /leads/export/excel    — download all leads (or filtered) as Excel (.xlsx)
  GET  /leads/export/template — download a blank CSV template for import
"""

import io
import csv
import uuid
import json
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, UploadFile, File, Query, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

router = APIRouter(prefix="/leads", tags=["Lead Import/Export"])

# ─── FIELD MAPPING ────────────────────────────────────────────────────────────
# Maps common CSV header variations → our DB column names
FIELD_ALIASES = {
    # name
    "name": "name", "full name": "name", "fullname": "name",
    "customer name": "name", "lead name": "name", "contact name": "name",
    # phone
    "phone": "phone", "mobile": "phone", "phone number": "phone",
    "mobile number": "phone", "contact number": "phone", "whatsapp": "phone",
    # email
    "email": "email", "email address": "email", "e-mail": "email",
    # module / course
    "module": "interested_module", "interested module": "interested_module",
    "course": "interested_module", "interested in": "interested_module",
    "interest": "interested_module", "program": "interested_module",
    # temperature
    "temperature": "temperature", "temp": "temperature", "lead temp": "temperature",
    # stage
    "stage": "lead_stage", "lead stage": "lead_stage", "status": "lead_stage",
    # source
    "source": "source", "lead source": "source", "channel": "source",
    # location
    "location": "location", "city": "location", "country": "location",
    # notes
    "notes": "notes", "note": "notes", "remarks": "notes", "comments": "notes",
    # experience
    "experience": "experience", "years of experience": "experience",
    # education
    "education": "education", "qualification": "education",
}

EXPORTABLE_FIELDS = [
    "id", "name", "phone", "email", "interested_module", "temperature",
    "lead_stage", "is_qualified", "source", "location", "experience",
    "education", "notes", "sender_id", "created_at", "updated_at",
]

TEMPLATE_FIELDS = [
    "name", "phone", "email", "interested_module",
    "temperature", "lead_stage", "source", "location", "notes",
]

VALID_TEMPERATURES = {"hot", "warm", "cold", ""}
VALID_STAGES = {"new", "qualified", "contacted", "demo_scheduled", "enrolled", "lost", ""}


def _supabase():
    from src.memory import supabase
    return supabase


def normalize_header(h: str) -> str:
    return h.strip().lower().replace("_", " ")


def map_headers(headers: list[str]) -> dict:
    """Returns {csv_col: db_field} mapping."""
    mapping = {}
    for h in headers:
        norm = normalize_header(h)
        if norm in FIELD_ALIASES:
            mapping[h] = FIELD_ALIASES[norm]
    return mapping


def parse_csv_bytes(content: bytes) -> tuple[list[str], list[dict]]:
    """Parse CSV bytes → (headers, rows)."""
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = [dict(row) for row in reader]
    return list(headers), rows


def parse_excel_bytes(content: bytes) -> tuple[list[str], list[dict]]:
    """Parse Excel bytes → (headers, rows). Requires openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = iter(ws.rows)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], []
    headers = [str(cell.value or "").strip() for cell in header_row]
    rows = []
    for row in rows_iter:
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(cell.value) if cell.value is not None else ""
        rows.append(row_dict)
    return headers, rows


def build_lead_row(csv_row: dict, mapping: dict, business_id: Optional[str] = None) -> dict:
    """Convert a CSV row dict into a leads table row dict using the mapping."""
    lead = {
        "sender_id": f"import_{uuid.uuid4().hex[:12]}",
        "source": "csv_import",
        "temperature": "cold",
        "lead_stage": "new",
        "is_qualified": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if business_id:
        lead["business_id"] = business_id

    for csv_col, db_field in mapping.items():
        val = (csv_row.get(csv_col) or "").strip()
        if not val:
            continue
        if db_field == "temperature":
            val = val.lower()
            if val not in VALID_TEMPERATURES:
                val = "cold"
        elif db_field == "lead_stage":
            val = val.lower().replace(" ", "_")
            if val not in VALID_STAGES:
                val = "new"
        elif db_field == "phone":
            # Normalise phone — strip spaces, dashes
            val = val.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
        lead[db_field] = val

    return lead


# ─── PREVIEW ENDPOINT ─────────────────────────────────────────────────────────

@router.post("/import/preview")
async def preview_import(file: UploadFile = File(...)):
    """
    Parse uploaded CSV or Excel file.
    Returns: detected headers, suggested column mapping, first 10 preview rows,
             total row count, and any validation warnings.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            headers, rows = parse_excel_bytes(content)
        else:
            headers, rows = parse_csv_bytes(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has no data rows.")

    mapping = map_headers(headers)

    # Validation warnings
    warnings = []
    if "name" not in mapping.values():
        warnings.append("No 'Name' column detected — leads will be imported without names.")
    if "phone" not in mapping.values() and "email" not in mapping.values():
        warnings.append("No 'Phone' or 'Email' column detected — leads will not be contactable.")

    # Build preview rows (first 10)
    preview = []
    for row in rows[:10]:
        built = build_lead_row(row, mapping)
        preview.append({
            "name": built.get("name", ""),
            "phone": built.get("phone", ""),
            "email": built.get("email", ""),
            "interested_module": built.get("interested_module", ""),
            "temperature": built.get("temperature", "cold"),
            "lead_stage": built.get("lead_stage", "new"),
            "source": built.get("source", "csv_import"),
            "notes": built.get("notes", ""),
        })

    return {
        "status": "success",
        "total_rows": len(rows),
        "headers": headers,
        "mapping": mapping,
        "preview": preview,
        "warnings": warnings,
        "unmapped_columns": [h for h in headers if h not in mapping],
    }


# ─── CONFIRM IMPORT ───────────────────────────────────────────────────────────

class ImportConfirmRequest(BaseModel):
    rows: list[dict]          # Already-built lead rows from frontend
    skip_duplicates: bool = True
    business_id: Optional[str] = None


@router.post("/import/confirm")
async def confirm_import(data: ImportConfirmRequest):
    """
    Insert the provided lead rows into the database.
    Checks for duplicates by phone/email if skip_duplicates=True.
    """
    try:
        supabase = _supabase()
        inserted = 0
        skipped = 0
        errors = []

        # Build sets of existing phones/emails for duplicate check
        existing_phones = set()
        existing_emails = set()
        if data.skip_duplicates:
            phone_res = supabase.table("leads").select("phone").not_.is_("phone", "null").execute()
            email_res = supabase.table("leads").select("email").not_.is_("email", "null").execute()
            existing_phones = {r["phone"] for r in (phone_res.data or []) if r.get("phone")}
            existing_emails = {r["email"] for r in (email_res.data or []) if r.get("email")}

        for i, row in enumerate(data.rows):
            try:
                # Duplicate check
                if data.skip_duplicates:
                    phone = row.get("phone", "")
                    email = row.get("email", "")
                    if phone and phone in existing_phones:
                        skipped += 1
                        continue
                    if email and email in existing_emails:
                        skipped += 1
                        continue

                # Ensure required fields
                row.setdefault("sender_id", f"import_{uuid.uuid4().hex[:12]}")
                row.setdefault("source", "csv_import")
                row.setdefault("temperature", "cold")
                row.setdefault("lead_stage", "new")
                row.setdefault("is_qualified", False)
                row.setdefault("created_at", datetime.now(timezone.utc).isoformat())
                row.setdefault("updated_at", datetime.now(timezone.utc).isoformat())
                if data.business_id:
                    row["business_id"] = data.business_id

                # Remove any keys not in leads table to avoid errors
                allowed_keys = {
                    "sender_id", "name", "phone", "email", "interested_module",
                    "temperature", "lead_stage", "is_qualified", "source",
                    "location", "experience", "education", "notes",
                    "created_at", "updated_at", "business_id",
                }
                clean_row = {k: v for k, v in row.items() if k in allowed_keys}

                supabase.table("leads").insert(clean_row).execute()
                inserted += 1

                # Track for duplicate prevention within same batch
                if row.get("phone"):
                    existing_phones.add(row["phone"])
                if row.get("email"):
                    existing_emails.add(row["email"])

            except Exception as e:
                errors.append(f"Row {i+1}: {str(e)}")

        return {
            "status": "success",
            "inserted": inserted,
            "skipped": skipped,
            "errors": errors[:10],  # Return first 10 errors max
            "message": f"Import complete — {inserted} leads added, {skipped} duplicates skipped."
        }

    except Exception as e:
        print(f"IMPORT ERROR: {e}", flush=True)
        raise HTTPException(status_code=500, detail=str(e))


# ─── EXPORT CSV ───────────────────────────────────────────────────────────────

@router.get("/export/csv")
async def export_leads_csv(
    temperature: Optional[str] = Query(None),
    lead_stage: Optional[str] = Query(None),
    is_qualified: Optional[bool] = Query(None),
    limit: int = Query(5000, le=10000),
):
    """Export leads as a CSV file download."""
    try:
        supabase = _supabase()
        query = supabase.table("leads").select(",".join(EXPORTABLE_FIELDS)).order("created_at", desc=True).limit(limit)

        if temperature:
            query = query.eq("temperature", temperature)
        if lead_stage:
            query = query.eq("lead_stage", lead_stage)
        if is_qualified is not None:
            query = query.eq("is_qualified", is_qualified)

        result = query.execute()
        leads = result.data or []

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=EXPORTABLE_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for lead in leads:
            writer.writerow({f: lead.get(f, "") for f in EXPORTABLE_FIELDS})

        output.seek(0)
        filename = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────

@router.get("/export/excel")
async def export_leads_excel(
    temperature: Optional[str] = Query(None),
    lead_stage: Optional[str] = Query(None),
    is_qualified: Optional[bool] = Query(None),
    limit: int = Query(5000, le=10000),
):
    """Export leads as an Excel (.xlsx) file download."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        supabase = _supabase()
        query = supabase.table("leads").select(",".join(EXPORTABLE_FIELDS)).order("created_at", desc=True).limit(limit)

        if temperature:
            query = query.eq("temperature", temperature)
        if lead_stage:
            query = query.eq("lead_stage", lead_stage)
        if is_qualified is not None:
            query = query.eq("is_qualified", is_qualified)

        result = query.execute()
        leads = result.data or []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        # Header row styling
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col_idx, field in enumerate(EXPORTABLE_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(15, len(field) + 4)

        # Data rows
        for row_idx, lead in enumerate(leads, start=2):
            for col_idx, field in enumerate(EXPORTABLE_FIELDS, start=1):
                val = lead.get(field, "")
                if isinstance(val, bool):
                    val = "Yes" if val else "No"
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Freeze header row
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── DOWNLOAD TEMPLATE ────────────────────────────────────────────────────────

@router.get("/export/template")
async def download_import_template():
    """Download a blank CSV template with the correct column headers for import."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=TEMPLATE_FIELDS)
    writer.writeheader()
    # Add one sample row so users understand the format
    writer.writerow({
        "name": "John Smith",
        "phone": "+919876543210",
        "email": "john@example.com",
        "interested_module": "SAP FICO",
        "temperature": "warm",
        "lead_stage": "new",
        "source": "Instagram",
        "location": "Mumbai",
        "notes": "Interested in weekend batch",
    })
    output.seek(0)

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=lead_import_template.csv"}
    )
